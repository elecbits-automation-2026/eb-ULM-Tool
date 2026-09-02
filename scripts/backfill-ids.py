"""Legacy → SOP v2.0 identity map for the artefact audit.

Every board in the audit gets a new-scheme identity. Serials run
INDEPENDENTLY per family (a board's PCB number says nothing about its
firmware number) and the Master tab is what joins them — exactly the
v2.0 doctrine: an identifier names a thing, the register relates them.

Rules applied, each traceable to the SOP:
  PCB  every board                                   EB-PCB-26-nnnn
  BOM  every board, as-designed revision             EB-PCB-26-nnnn-BOM-001  (derived, per parent)
  FW   only boards with firmware evidence            EB-FW-26-nnnn   (Law 7: no GW/SS marker)
  ED   one per project with enclosure evidence,      EB-ED-26-nnnn   (the register files
       one per orphan board with enclosure evidence                   enclosures against a project)
  P    one per distinct legacy project id            EB-P-26-nnnn
Boards with no legacy project keep an empty Project ID: a project comes
from a won deal (Law 10) and cannot be invented here.

The old Board ID is never destroyed — it lands in Name / Alias and in
Legacy SKU Code, which is where descriptive, changeable data belongs.
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import re, sys

SRC = "/root/.claude/uploads/8dd0231e-23c0-58ed-87e3-c214dcbc622a/22672866-EbArtefact_Presence_Audit_v2_with_ProjectID.xlsx"
YY = "26"
# First free serial per family. Raise these to reserve a block after whatever
# the live master register already holds.
START = {"P": 1, "PCB": 1, "FW": 1, "ED": 1}

ws = load_workbook(SRC, data_only=True)["Audit 01-Sep 2233"]
hdr = [("" if c is None else str(c).strip()) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
col = {h: i for i, h in enumerate(hdr)}
FW_C = [i for i, h in enumerate(hdr) if h.startswith("FW ") and not h.endswith("- files")]
EN_C = [i for i, h in enumerate(hdr) if h.startswith("EN ") and not h.endswith("- files")]
HW_C = [i for i, h in enumerate(hdr) if h.startswith("HW ") and not h.endswith("- files")]
present = lambda r, cs: any("PRESENT" in ("" if r[i] is None else str(r[i]).upper()) for i in cs)

def cls(board):                      # the old scheme's class marker, kept as data only
    m = re.search(r"-(GW|SS)(?:-\d+)?$", board.strip().upper())
    return {"GW": "Gateway", "SS": "Sensor Node"}.get(m.group(1), "Other") if m else "Other"
def platform(board):                 # first token was the MCU in the old naming
    return board.strip().split("-")[0].upper()

boards = []
for r in rows:
    b = str(r[0]).strip()
    boards.append({
        "board": b, "legacy_project": (str(r[1]).strip() if r[1] else ""),
        "folder": str(r[2] or "").strip(), "owner": str(r[col.get("Folder Owner", 32)] or "").strip(),
        "hw": present(r, HW_C), "fw": present(r, FW_C), "en": present(r, EN_C),
        "platform": platform(b), "class": cls(b),
    })
boards.sort(key=lambda x: x["board"].upper())          # stable, reproducible order

def mk(fam, n): return f"EB-{fam}-{YY}-{n:04d}"

# projects — one per distinct legacy id, in first-appearance order
proj_ids, seen = {}, []
for b in boards:
    if b["legacy_project"] and b["legacy_project"] not in proj_ids:
        proj_ids[b["legacy_project"]] = mk("P", START["P"] + len(seen)); seen.append(b["legacy_project"])

n_pcb = n_fw = n_ed = 0
ed_by_project = {}
for b in boards:
    n_pcb += 1
    b["pcb"] = mk("PCB", START["PCB"] + n_pcb - 1)
    b["bom"] = b["pcb"] + "-BOM-001"
    b["project"] = proj_ids.get(b["legacy_project"], "")
    if b["fw"]:
        n_fw += 1; b["fw_id"] = mk("FW", START["FW"] + n_fw - 1)
    else:
        b["fw_id"] = ""
    if b["en"]:
        key = b["project"] or f"__board__{b['pcb']}"      # one enclosure serves a project
        if key not in ed_by_project:
            n_ed += 1; ed_by_project[key] = mk("ED", START["ED"] + n_ed - 1)
        b["ed_id"] = ed_by_project[key]
    else:
        b["ed_id"] = ""

wb = Workbook(); wb.remove(wb.active)
H1 = Font(bold=True, color="FFFFFF"); FILL = PatternFill("solid", fgColor="1F3864")
def sheet(name, headers, data):
    s = wb.create_sheet(name)
    s.append(headers)
    for c in s[1]: c.font = H1; c.fill = FILL; c.alignment = Alignment(vertical="center")
    for row in data: s.append(row)
    for i, h in enumerate(headers, 1):
        s.column_dimensions[s.cell(1, i).column_letter].width = max(12, min(34, len(h) + 6))
    s.freeze_panes = "A2"
    return s

sheet("Map", ["Board ID (legacy)", "Legacy Project ID", "EB Project ID", "EB PCB ID", "EB BOM ID",
              "EB FW ID", "EB Enclosure ID", "Platform", "Class", "HW?", "FW?", "Enclosure?", "Folder name as stored"],
      [[b["board"], b["legacy_project"], b["project"], b["pcb"], b["bom"], b["fw_id"], b["ed_id"],
        b["platform"], b["class"], "yes" if b["hw"] else "", "yes" if b["fw"] else "",
        "yes" if b["en"] else "", b["folder"]] for b in boards])

sheet("Projects", ["Project ID", "Source Deal ID", "Client ID", "Project Name", "Kind", "Status",
                   "Project Manager", "Start Date", "Drive Folder Link", "Date Added", "Added By", "Notes"],
      [[pid, "", "", legacy, "", "Active", "", "", "", "", "backfill",
        f"Legacy project {legacy} — {sum(1 for b in boards if b['project']==pid)} board(s). Source deal and client to be attached."]
       for legacy, pid in ((l, proj_ids[l]) for l in seen)])

sheet("PCB", ["PCB ID", "Project ID", "Name / Alias", "Drive Folder Link", "Legacy SKU Code",
              "Silkscreen Marking", "Platform", "Class", "Version", "Status", "Date Added", "Added By", "Notes"],
      [[b["pcb"], b["project"], b["board"], "", b["board"], f"{b['pcb']} V1", b["platform"], b["class"],
        "V1", "Active", "", "backfill", "Legacy board carried over from the artefact audit"] for b in boards])

sheet("BOM", ["BOM ID", "PCB ID", "Revision Reason", "Line Count", "Costed?", "Cost per Unit",
              "Costed On", "Status", "Date Added", "Added By", "Notes"],
      [[b["bom"], b["pcb"], "As designed", "", "", "", "", "Active", "", "backfill",
        "BOM-001 is always the as-designed revision"] for b in boards])

sheet("FW", ["FW ID", "PCB ID", "Project ID", "Platform", "Latest Version (Git tag)", "Repo",
             "Drive Folder Link", "Status", "Date Added", "Added By", "Notes"],
      [[b["fw_id"], b["pcb"], b["project"], b["platform"], "", f"fw-product-eb-fw-{YY}-{b['fw_id'][-4:]}",
        "", "Active", "", "backfill", "Firmware artefacts found in the audit"] for b in boards if b["fw_id"]])

seen_ed = {}
ed_rows = []
for b in boards:
    if b["ed_id"] and b["ed_id"] not in seen_ed:
        seen_ed[b["ed_id"]] = True
        ed_rows.append([b["ed_id"], b["project"], f"Enclosure for {b['board']}", "", "", "V1", "Active",
                        "", "backfill", "Enclosure artefacts found in the audit"])
sheet("Enclosure", ["Enclosure ID", "Project ID", "Name", "Drive Folder Link", "Material", "Version",
                    "Status", "Date Added", "Added By", "Notes"], ed_rows)

sheet("Master", ["Client ID", "Deal ID", "Project ID", "PCB ID", "BOM ID", "FW ID", "Enclosure ID",
                 "MFG ID", "Client Name (auto)", "Project Name (auto)", "Rule", "Notes"],
      [["", "", b["project"], b["pcb"], b["bom"], b["fw_id"], b["ed_id"], "", "", b["legacy_project"],
        "1.0", f"Backfill of {b['board']}"] for b in boards])

out = "EbID_Backfill_v2.xlsx"
wb.save(out)
print(f"boards={len(boards)}  projects={len(proj_ids)}  pcb={n_pcb}  bom={n_pcb}  fw={n_fw}  ed={n_ed}")
print("last serials:", mk("P", START['P']+len(proj_ids)-1), mk("PCB", START['PCB']+n_pcb-1), mk("FW", START['FW']+n_fw-1), mk("ED", START['ED']+n_ed-1))
print("wrote", out)
