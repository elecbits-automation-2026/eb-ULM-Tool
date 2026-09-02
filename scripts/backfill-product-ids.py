"""EVSO / Pro-connect / Repeater — four boards per version, given v2.0 ids.

The version tracker lists each product's versions as separate legacy
projects. Five of those already earned an EB-P number in the artefact-audit
backfill, so they are REUSED here — Law 1: nothing is ever renumbered. Only
the two versions the audit never saw are minted fresh.

Serials continue after the audit run rather than restarting, because the two
backfills write into the same register: an issued number is spent.

Each version has four boards, and each board is a thing in its own right:
  PCB  EB-PCB-YY-nnnn            one per board per version
  BOM  EB-PCB-YY-nnnn-BOM-001    the as-designed revision of that board
  FW   EB-FW-YY-nnnn             the board's firmware identity

A respun board in the next version is a NEW board identifier, not a version
of the old one — that is rule 5.0/6.0, and it is why V1's mainboard and V3's
mainboard cannot share a number.
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

VT = "/root/.claude/uploads/8dd0231e-23c0-58ed-87e3-c214dcbc622a/1c699a38-Version_trackingEVSO_Proconnect_Repeater_1.xlsx"
YY = "26"
# Continue after the artefact-audit backfill (P→0038, PCB→0153, FW→0059).
START = {"P": 39, "PCB": 154, "FW": 60}
# EB-P numbers the audit already issued to these legacy projects.
REUSE_P = {
    "Eb-21-EL-287-01-1453": "EB-P-26-0028", "Eb-21-EL-287-01-1628": "EB-P-26-0008",
    "Eb-21-EL-287-01-1452": "EB-P-26-0029", "Eb-21-EL-287-01-1629": "EB-P-26-0009",
    "Eb-21-EL-287-01-1579": "EB-P-26-0007",
}
BOARDS = {
    "EVSO (Outdoor)":       ["Mainboard", "Daughterboard", "HMI", "LED"],
    "Repeater":             ["Mainboard", "Daughterboard", "HMI", "LED"],
    "Pro-connect (Indoor)": ["HMI", "Power", "Left", "Right"],
}
# ONE firmware identity per product version, not per board: the unit runs a
# single build across its boards. The FW row names its host board, and the
# Master rows tie that firmware to every board of the version.
FW_HOST = {"EVSO (Outdoor)": "Mainboard", "Repeater": "Mainboard", "Pro-connect (Indoor)": "HMI"}
# Boards that carry their own bill of material. Pro-connect's Left and Right
# boards are covered by the HMI and Power BOMs, so they get no BOM of their own.
BOM_BOARDS = {"EVSO (Outdoor)": {"Mainboard", "Daughterboard", "HMI", "LED"},
              "Repeater":       {"Mainboard", "Daughterboard", "HMI", "LED"},
              "Pro-connect (Indoor)": {"HMI", "Power"}}
CLASS = {"Mainboard": "Gateway", "Daughterboard": "Controller", "HMI": "Controller",
         "LED": "Power", "Power": "Power", "Left": "Sensor Node", "Right": "Sensor Node"}

ws = load_workbook(VT, data_only=True)["Version tracking"]
versions, product = [], ""
for r in ws.iter_rows(min_row=2, values_only=True):
    if not r[1]: continue
    product = (r[0] or product)
    versions.append({"product": product.strip(), "version": str(r[1]).strip(),
                     "legacy_project": str(r[2] or "").strip(), "legacy_product": str(r[5] or "").strip()})

mk = lambda fam, n: f"EB-{fam}-{YY}-{n:04d}"
nP = nPcb = nFw = 0
rows = []
for v in versions:
    lp = v["legacy_project"]
    if lp in REUSE_P:
        v["project"], v["project_src"] = REUSE_P[lp], "reused from the audit backfill"
    else:
        v["project"], v["project_src"] = mk("P", START["P"] + nP), "newly issued"; nP += 1
    # One firmware for the whole version, issued once and shared by its boards.
    v["fw"] = mk("FW", START["FW"] + nFw); nFw += 1
    for board in BOARDS[v["product"]]:
        pcb = mk("PCB", START["PCB"] + nPcb); nPcb += 1
        has_bom = board in BOM_BOARDS[v["product"]]
        rows.append({
            "product": v["product"], "version": v["version"], "board": board,
            "project": v["project"], "project_src": v["project_src"],
            "legacy_project": lp, "legacy_product": v["legacy_product"],
            "pcb": pcb, "bom": (pcb + "-BOM-001") if has_bom else "",
            "fw": v["fw"], "fw_host": board == FW_HOST[v["product"]],
            "bom_note": "" if has_bom else "covered by the HMI and Power BOMs",
            "cls": CLASS[board],
            "alias": f"{v['product']} {v['version']} — {board}",
        })

wb = Workbook(); wb.remove(wb.active)
H = Font(bold=True, color="FFFFFF"); F = PatternFill("solid", fgColor="1F3864")
def sheet(name, headers, data):
    s = wb.create_sheet(name); s.append(headers)
    for c in s[1]: c.font = H; c.fill = F; c.alignment = Alignment(vertical="center")
    for d in data: s.append(d)
    for i, h in enumerate(headers, 1):
        s.column_dimensions[s.cell(1, i).column_letter].width = max(12, min(38, len(h) + 6))
    s.freeze_panes = "A2"

sheet("Map", ["Product", "Version", "Board", "EB Project ID", "EB PCB ID", "EB BOM ID", "EB FW ID",
              "FW host board", "BOM note", "Class", "Legacy Project ID", "Legacy Product ID", "Project ID source"],
      [[r["product"], r["version"], r["board"], r["project"], r["pcb"], r["bom"], r["fw"],
        "yes" if r["fw_host"] else "", r["bom_note"], r["cls"], r["legacy_project"], r["legacy_product"],
        r["project_src"]] for r in rows])

sheet("PCB", ["PCB ID", "Project ID", "Name / Alias", "Drive Folder Link", "Legacy SKU Code",
              "Silkscreen Marking", "Platform", "Class", "Version", "Status", "Date Added", "Added By", "Notes"],
      [[r["pcb"], r["project"], r["alias"], "", r["legacy_product"], f"{r['pcb']} V1", "", r["cls"],
        r["version"], "Active", "", "backfill", f"{r['board']} of {r['product']} {r['version']}"] for r in rows])

sheet("BOM", ["BOM ID", "PCB ID", "Revision Reason", "Line Count", "Costed?", "Cost per Unit",
              "Costed On", "Status", "Date Added", "Added By", "Notes"],
      [[r["bom"], r["pcb"], "As designed", "", "", "", "", "Active", "", "backfill",
        f"BOM-001 is always the as-designed revision — {r['board']} of {r['product']} {r['version']}"]
       for r in rows if r["bom"]])

sheet("FW", ["FW ID", "PCB ID", "Project ID", "Platform", "Latest Version (Git tag)", "Repo",
             "Drive Folder Link", "Status", "Date Added", "Added By", "Notes"],
      [[r["fw"], r["pcb"], r["project"], "", "", f"fw-product-eb-fw-{YY}-{r['fw'][-4:]}", "",
        "Active", "", "backfill",
        f"One firmware for {r['product']} {r['version']} — hosted on the {r['board']}, running across all "
        + str(len(BOARDS[r['product']])) + " boards of the version"]
       for r in rows if r["fw_host"]])

newp = [v for v in versions if v["project_src"] == "newly issued"]
sheet("Projects (new only)", ["Project ID", "Source Deal ID", "Client ID", "Project Name", "Kind", "Status",
                              "Project Manager", "Start Date", "Drive Folder Link", "Date Added", "Added By", "Notes"],
      [[v["project"], "", "", f"{v['product']} {v['version']}", "RND+MFG", "Active", "", "", "", "", "backfill",
        f"Legacy project {v['legacy_project']}"] for v in newp])

sheet("Master", ["Client ID", "Deal ID", "Project ID", "PCB ID", "BOM ID", "FW ID", "Enclosure ID",
                 "MFG ID", "Client Name (auto)", "Project Name (auto)", "Rule", "Notes"],
      [["", "", r["project"], r["pcb"], r["bom"], r["fw"], "", "", "",
        f"{r['product']} {r['version']}", "1.0",
        f"{r['board']} — one firmware across the version" + (" · no BOM of its own" if not r["bom"] else "")]
       for r in rows])

wb.save("EbID_EVSO_Proconnect_Repeater.xlsx")
nBom = sum(1 for r in rows if r["bom"])
print(f"versions={len(versions)}  boards={len(rows)}  new projects={nP}  pcb={nPcb}  bom={nBom}  fw={nFw}")
print(f"ranges: PCB {rows[0]['pcb']}..{rows[-1]['pcb']}   FW {rows[0]['fw']}..{rows[-1]['fw']}")
print("boards without their own BOM:", sum(1 for r in rows if not r["bom"]))
