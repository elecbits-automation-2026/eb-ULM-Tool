"""The audit workbook, returned with its v2.0 identities filled in.

Two allocations meet here and must not collide:
  · the 153 audited boards            EB-PCB-26-0001..0153
  · the 28 product boards (EVSO /     EB-PCB-26-0154..0181
    Pro-connect / Repeater versions)
Both keep exactly the numbers already reviewed — nothing is renumbered.

Five audited rows are the OLD gateway-board names of those same products
(ES3C5-STE5-LK306-GW-101 and friends). They named a product, not one of its
four boards, so their rows are marked Superseded and point at the board-level
identities that replace them. Law 1: a row is retired, never deleted.
"""
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import re

AUDIT = "/root/.claude/uploads/8dd0231e-23c0-58ed-87e3-c214dcbc622a/ef57b7e5-EbArtefact_Presence_Audit_v2_with_ProjectID_1.xlsx"
VT    = "/root/.claude/uploads/8dd0231e-23c0-58ed-87e3-c214dcbc622a/1c699a38-Version_trackingEVSO_Proconnect_Repeater_1.xlsx"
YY = "26"
mk = lambda f, n: f"EB-{f}-{YY}-{n:04d}"

# ── 1. the audited boards ──────────────────────────────────────────────────
src = load_workbook(AUDIT, data_only=True)
aud = src["Audit 01-Sep 2233"]
hdr = [("" if c is None else str(c).strip()) for c in next(aud.iter_rows(min_row=1, max_row=1, values_only=True))]
raw = [list(r) for r in aud.iter_rows(min_row=2, values_only=True) if r[0]]
FW_C = [i for i, h in enumerate(hdr) if h.startswith("FW ") and not h.endswith("- files")]
EN_C = [i for i, h in enumerate(hdr) if h.startswith("EN ") and not h.endswith("- files")]
HW_C = [i for i, h in enumerate(hdr) if h.startswith("HW ") and not h.endswith("- files")]
seen = lambda r, cs: any("PRESENT" in str(r[i] or "").upper() for i in cs)
cls  = lambda b: "Gateway" if re.search(r"-GW(-\d+)?$", b.upper()) else ("Sensor Node" if re.search(r"-SS(-\d+)?$", b.upper()) else "Other")

boards = [{"board": str(r[0]).strip(), "legacy_project": str(r[1] or "").strip(),
           "hw": seen(r, HW_C), "fw": seen(r, FW_C), "en": seen(r, EN_C),
           "platform": str(r[0]).strip().split("-")[0].upper(), "cls": cls(str(r[0]).strip()),
           "raw": r} for r in raw]
order = sorted(range(len(boards)), key=lambda i: boards[i]["board"].upper())   # allocation order

proj, seq = {}, []
for i in order:
    lp = boards[i]["legacy_project"]
    if lp and lp not in proj:
        proj[lp] = mk("P", 1 + len(seq)); seq.append(lp)
nP, nPcb, nFw, nEd, ed_for = len(proj), 0, 0, 0, {}
for i in order:
    b = boards[i]
    nPcb += 1; b["pcb"] = mk("PCB", nPcb); b["bom"] = b["pcb"] + "-BOM-001"
    b["project"] = proj.get(b["legacy_project"], "")
    b["fw_id"] = ""
    if b["fw"]:
        nFw += 1; b["fw_id"] = mk("FW", nFw)
    b["ed_id"] = ""
    if b["en"]:
        key = b["project"] or "board:" + b["pcb"]
        if key not in ed_for:
            nEd += 1; ed_for[key] = mk("ED", nEd)
        b["ed_id"] = ed_for[key]

# ── 2. the product boards, continuing the same runs ────────────────────────
REUSE = {"Eb-21-EL-287-01-1453": "", "Eb-21-EL-287-01-1628": "", "Eb-21-EL-287-01-1452": "",
         "Eb-21-EL-287-01-1629": "", "Eb-21-EL-287-01-1579": ""}
for lp in REUSE: REUSE[lp] = proj.get(lp, "")          # reuse whatever the audit issued
BOARDS = {"EVSO (Outdoor)": ["Mainboard", "Daughterboard", "HMI", "LED"],
          "Repeater": ["Mainboard", "Daughterboard", "HMI", "LED"],
          "Pro-connect (Indoor)": ["HMI", "Power", "Left", "Right"]}
HOST = {"EVSO (Outdoor)": "Mainboard", "Repeater": "Mainboard", "Pro-connect (Indoor)": "HMI"}
BOM_ONLY = {"EVSO (Outdoor)": None, "Repeater": None, "Pro-connect (Indoor)": ["HMI", "Power"]}
PCLASS = {"Mainboard": "Gateway", "Daughterboard": "Controller", "HMI": "Controller",
          "LED": "Power", "Power": "Power", "Left": "Sensor Node", "Right": "Sensor Node"}

vt = load_workbook(VT, data_only=True)["Version tracking"]
versions, prod = [], ""
for r in vt.iter_rows(min_row=2, values_only=True):
    if not r[1]: continue
    prod = (r[0] or prod)
    versions.append({"product": prod.strip(), "version": str(r[1]).strip(),
                     "legacy": str(r[2] or "").strip(), "legacy_product": str(r[5] or "").strip()})
prods = []
for v in versions:
    v["project"] = REUSE.get(v["legacy"]) or mk("P", 1 + nP); 
    if not REUSE.get(v["legacy"]): nP += 1
    v["fw"] = mk("FW", nFw + 1); nFw += 1
    for board in BOARDS[v["product"]]:
        nPcb += 1
        bo = BOM_ONLY[v["product"]]
        has = bo is None or board in bo
        pcb = mk("PCB", nPcb)
        prods.append({"product": v["product"], "version": v["version"], "board": board,
                      "project": v["project"], "pcb": pcb, "bom": pcb + "-BOM-001" if has else "",
                      "fw": v["fw"], "host": board == HOST[v["product"]], "cls": PCLASS[board],
                      "legacy": v["legacy"], "legacy_product": v["legacy_product"],
                      "alias": f"{v['product']} {v['version']} — {board}"})

# ── 3. the overlap: audited rows that named a product, not a board ─────────
SUPERSEDED = {}
for v in versions:
    if not v["legacy_product"]: continue
    for b in boards:
        if b["board"].upper() == v["legacy_product"].upper():
            rng = [p["pcb"] for p in prods if p["product"] == v["product"] and p["version"] == v["version"]]
            SUPERSEDED[b["pcb"]] = (f"{v['product']} {v['version']}", rng)
            b["superseded"] = f"Product-level entry — replaced by the four board identities {rng[0]}..{rng[-1]}"

# ── 4. write the workbook ──────────────────────────────────────────────────
wb = Workbook(); wb.remove(wb.active)
H = Font(bold=True, color="FFFFFF"); F = PatternFill("solid", fgColor="1F3864")
GREY = Font(color="808080", italic=True)
def sheet(name, head, rows, width_cap=34):
    s = wb.create_sheet(name); s.append(head)
    for c in s[1]: c.font = H; c.fill = F; c.alignment = Alignment(vertical="center")
    for r in rows: s.append(r)
    for i, h in enumerate(head, 1):
        s.column_dimensions[get_column_letter(i)].width = max(11, min(width_cap, len(str(h)) + 5))
    s.freeze_panes = "A2"
    return s

# 4a. the audit tab, with its identities filled in
NEWCOLS = ["EB Project ID", "EB PCB ID", "EB BOM ID", "EB FW ID", "EB Enclosure ID", "ID note"]
a = sheet("Audit 01-Sep 2233", hdr + NEWCOLS,
          [list(b["raw"]) + [b["project"], b["pcb"], b["bom"], b["fw_id"], b["ed_id"],
                             b.get("superseded", "")] for b in boards], width_cap=26)
for i, b in enumerate(boards, start=2):
    if b.get("superseded"):
        for c in range(len(hdr) + 1, len(hdr) + len(NEWCOLS) + 1): a.cell(i, c).font = GREY

# 4b. the product breakdown
sheet("EB Product IDs", ["Product", "Version", "Board", "EB Project ID", "EB PCB ID", "EB BOM ID",
                         "EB FW ID", "FW host", "BOM note", "Class", "Legacy Project ID", "Legacy Product ID"],
      [[p["product"], p["version"], p["board"], p["project"], p["pcb"], p["bom"], p["fw"],
        "yes" if p["host"] else "", "" if p["bom"] else "covered by the HMI and Power BOMs",
        p["cls"], p["legacy"], p["legacy_product"]] for p in prods])

# 4c. register-ready tabs — both allocations, in one run each
sup = lambda pcb: "Superseded" if pcb in SUPERSEDED else "Active"
sheet("PCB", ["PCB ID", "Project ID", "Name / Alias", "Drive Folder Link", "Legacy SKU Code",
              "Silkscreen Marking", "Platform", "Class", "Version", "Status", "Date Added", "Added By", "Notes"],
      [[b["pcb"], b["project"], b["board"], "", b["board"], b["pcb"] + " V1", b["platform"], b["cls"], "V1",
        sup(b["pcb"]), "", "backfill", b.get("superseded", "Legacy board carried over from the artefact audit")]
       for b in boards]
      + [[p["pcb"], p["project"], p["alias"], "", p["legacy_product"], p["pcb"] + " V1", "", p["cls"],
          p["version"], "Active", "", "backfill", f"{p['board']} of {p['product']} {p['version']}"] for p in prods])

sheet("BOM", ["BOM ID", "PCB ID", "Revision Reason", "Line Count", "Costed?", "Cost per Unit",
              "Costed On", "Status", "Date Added", "Added By", "Notes"],
      [[b["bom"], b["pcb"], "As designed", "", "", "", "", sup(b["pcb"]), "", "backfill",
        "BOM-001 is always the as-designed revision"] for b in boards if b["pcb"] not in SUPERSEDED]
      + [[p["bom"], p["pcb"], "As designed", "", "", "", "", "Active", "", "backfill",
          f"As-designed revision of the {p['board']}"] for p in prods if p["bom"]])

sheet("FW", ["FW ID", "PCB ID", "Project ID", "Platform", "Latest Version (Git tag)", "Repo",
             "Drive Folder Link", "Status", "Date Added", "Added By", "Notes"],
      [[b["fw_id"], b["pcb"], b["project"], b["platform"], "", f"fw-product-eb-fw-{YY}-{b['fw_id'][-4:]}", "",
        sup(b["pcb"]), "", "backfill",
        b.get("superseded", "Firmware artefacts found in the audit")] for b in boards if b["fw_id"]]
      + [[p["fw"], p["pcb"], p["project"], "", "", f"fw-product-eb-fw-{YY}-{p['fw'][-4:]}", "", "Active", "",
          "backfill", f"One firmware for {p['product']} {p['version']} — hosted on the {p['board']}, "
                      f"running across all {len(BOARDS[p['product']])} boards"] for p in prods if p["host"]])

edrows, done = [], set()
for b in boards:
    if b["ed_id"] and b["ed_id"] not in done:
        done.add(b["ed_id"])
        edrows.append([b["ed_id"], b["project"], f"Enclosure for {b['board']}", "", "", "V1", "Active",
                       "", "backfill", "Enclosure artefacts found in the audit"])
sheet("Enclosure", ["Enclosure ID", "Project ID", "Name", "Drive Folder Link", "Material", "Version",
                    "Status", "Date Added", "Added By", "Notes"], edrows)

pjrows = [[proj[lp], "", "", lp, "", "Active", "", "", "", "", "backfill",
           f"Legacy project {lp} — {sum(1 for b in boards if b['project']==proj[lp])} audited board(s)"] for lp in seq]
newv = [v for v in versions if not REUSE.get(v["legacy"])]
pjrows += [[v["project"], "", "", f"{v['product']} {v['version']}", "RND+MFG", "Active", "", "", "", "",
            "backfill", f"Legacy project {v['legacy']}"] for v in newv]
sheet("Projects", ["Project ID", "Source Deal ID", "Client ID", "Project Name", "Kind", "Status",
                   "Project Manager", "Start Date", "Drive Folder Link", "Date Added", "Added By", "Notes"], pjrows)

sheet("Master", ["Client ID", "Deal ID", "Project ID", "PCB ID", "BOM ID", "FW ID", "Enclosure ID", "MFG ID",
                 "Legacy Board ID", "Legacy Project ID", "Platform", "Class", "Rule", "Notes"],
      [["", "", b["project"], b["pcb"], "" if b["pcb"] in SUPERSEDED else b["bom"], b["fw_id"], b["ed_id"], "",
        b["board"], b["legacy_project"], b["platform"], b["cls"], "1.0",
        b.get("superseded", "audit backfill")] for b in boards]
      + [["", "", p["project"], p["pcb"], p["bom"], p["fw"], "", "", p["legacy_product"], p["legacy"], "",
          p["cls"], "1.0", f"product backfill — {p['product']} {p['version']} {p['board']}"] for p in prods])

# keep the owners tab as it came
own = src["Folder Owners"]
o = wb.create_sheet("Folder Owners")
for r in own.iter_rows(values_only=True):
    if any(v is not None for v in r): o.append(list(r))
for c in o[1]: c.font = H; c.fill = F

out = "EbArtefact_Audit_v2_with_EB_IDs.xlsx"
wb.save(out)
print(f"audited boards={len(boards)}  product boards={len(prods)}")
print(f"projects={len(pjrows)}  pcb={len(boards)+len(prods)}  fw={sum(1 for b in boards if b['fw_id'])+len([p for p in prods if p['host']])}  enclosures={len(edrows)}")
print(f"superseded product-level rows={len(SUPERSEDED)}: {', '.join(sorted(SUPERSEDED))}")
print("wrote", out)
